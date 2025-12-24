"""
검증용 엑셀 생성 (매칭 문항 포함)
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import Counter
import os

os.chdir('c:/Project/CJ_Culture')

print('=== 검증용 엑셀 생성 (매칭 문항 포함) ===')

# 데이터 로드
df = pd.read_pickle('all_df_hybrid.pkl')
print(f'총 문항: {len(df)}')

# 스타일 정의
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

cat_fills = {
    'OD': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'LD': PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid'),
    'MA': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'DD': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
}

wb = Workbook()

# ====== Sheet 1: 전체요약 ======
ws_summary = wb.active
ws_summary.title = '전체요약'

headers = ['대분류', '문항수', '분류됨', '분류율(%)', '클러스터수', '미분류']
for col, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row = 2
total_q, total_c, total_u = 0, 0, 0
for cat in ['OD', 'LD', 'MA', 'DD']:
    cat_df = df[df['대분류'] == cat]
    classified = len(cat_df[cat_df['중분류'] != '미분류'])
    unclassified = len(cat_df[cat_df['중분류'] == '미분류'])
    clusters = cat_df['cluster_id'].nunique()

    ws_summary.cell(row=row, column=1, value=cat).fill = cat_fills[cat]
    ws_summary.cell(row=row, column=2, value=len(cat_df))
    ws_summary.cell(row=row, column=3, value=classified)
    ws_summary.cell(row=row, column=4, value=round(classified/len(cat_df)*100, 1))
    ws_summary.cell(row=row, column=5, value=clusters)
    ws_summary.cell(row=row, column=6, value=unclassified)
    for c in range(1, 7):
        ws_summary.cell(row=row, column=c).border = thin_border
    total_q += len(cat_df)
    total_c += classified
    total_u += unclassified
    row += 1

ws_summary.cell(row=row, column=1, value='합계').font = Font(bold=True)
ws_summary.cell(row=row, column=2, value=total_q).font = Font(bold=True)
ws_summary.cell(row=row, column=3, value=total_c).font = Font(bold=True)
ws_summary.cell(row=row, column=4, value=round(total_c/total_q*100, 1)).font = Font(bold=True)
ws_summary.cell(row=row, column=6, value=total_u).font = Font(bold=True)
for c in range(1, 7):
    ws_summary.cell(row=row, column=c).border = thin_border

for i, w in enumerate([10, 10, 10, 12, 12, 10], 1):
    ws_summary.column_dimensions[chr(64+i)].width = w

print('Sheet 1: 전체요약 완료')

# ====== Sheet 2: 중분류별 요약 ======
ws_mid = wb.create_sheet(title='중분류별요약')
headers = ['중분류', '문항수', '소분류수', '클러스터수']
for col, h in enumerate(headers, 1):
    cell = ws_mid.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

mid_stats = df.groupby('중분류').agg({
    '문항': 'count',
    '소분류': 'nunique',
    'cluster_id': 'nunique'
}).reset_index()
mid_stats.columns = ['중분류', '문항수', '소분류수', '클러스터수']
mid_stats = mid_stats.sort_values('문항수', ascending=False)

row = 2
for _, r in mid_stats.iterrows():
    ws_mid.cell(row=row, column=1, value=r['중분류'])
    ws_mid.cell(row=row, column=2, value=r['문항수'])
    ws_mid.cell(row=row, column=3, value=r['소분류수'])
    ws_mid.cell(row=row, column=4, value=r['클러스터수'])
    for c in range(1, 5):
        ws_mid.cell(row=row, column=c).border = thin_border
    row += 1

ws_mid.column_dimensions['A'].width = 20
ws_mid.column_dimensions['B'].width = 10
ws_mid.column_dimensions['C'].width = 10
ws_mid.column_dimensions['D'].width = 12

print('Sheet 2: 중분류별요약 완료')

# ====== Sheet 3: 클러스터_매칭문항 ======
ws_match = wb.create_sheet(title='클러스터_매칭문항')
headers = ['대분류', '중분류', '소분류', 'Cluster_ID', '문항수', '대표문항', '매칭문항들']
for col, h in enumerate(headers, 1):
    cell = ws_match.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

row = 2
for cat in ['OD', 'LD', 'MA', 'DD']:
    cat_df = df[df['대분류'] == cat]
    for cluster_id in sorted(cat_df['cluster_id'].unique()):
        cluster_df = cat_df[cat_df['cluster_id'] == cluster_id]

        mid_counts = cluster_df['중분류'].value_counts()
        main_mid = mid_counts.index[0]

        small_counts = cluster_df['소분류'].value_counts()
        main_small = small_counts.index[0]

        rep_q = cluster_df.iloc[0]['문항']
        if len(rep_q) > 50:
            rep_q = rep_q[:50] + '...'

        match_qs = cluster_df['문항'].tolist()[:5]
        match_list = []
        for q in match_qs:
            if len(q) > 40:
                match_list.append(q[:40] + '...')
            else:
                match_list.append(q)
        match_str = ' | '.join(match_list)

        ws_match.cell(row=row, column=1, value=cat).fill = cat_fills[cat]
        ws_match.cell(row=row, column=2, value=main_mid)
        ws_match.cell(row=row, column=3, value=main_small)
        ws_match.cell(row=row, column=4, value=cluster_id)
        ws_match.cell(row=row, column=5, value=len(cluster_df))
        ws_match.cell(row=row, column=6, value=rep_q)
        ws_match.cell(row=row, column=7, value=match_str)
        for c in range(1, 8):
            ws_match.cell(row=row, column=c).border = thin_border
        row += 1

ws_match.column_dimensions['A'].width = 8
ws_match.column_dimensions['B'].width = 18
ws_match.column_dimensions['C'].width = 18
ws_match.column_dimensions['D'].width = 12
ws_match.column_dimensions['E'].width = 8
ws_match.column_dimensions['F'].width = 55
ws_match.column_dimensions['G'].width = 100

print(f'Sheet 3: 클러스터_매칭문항 완료 ({row-2}개 클러스터)')

# ====== Sheet 4: 전체문항상세 ======
ws_detail = wb.create_sheet(title='전체문항상세')
headers = ['대분류', '중분류', '소분류', 'Cluster_ID', '문항', '년도', '회사명']
for col, h in enumerate(headers, 1):
    cell = ws_detail.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

df_sorted = df.sort_values(['대분류', '중분류', '소분류', 'cluster_id'])
row = 2
for _, r in df_sorted.iterrows():
    ws_detail.cell(row=row, column=1, value=r['대분류']).fill = cat_fills.get(r['대분류'], PatternFill())
    ws_detail.cell(row=row, column=2, value=r['중분류'])
    ws_detail.cell(row=row, column=3, value=r['소분류'])
    ws_detail.cell(row=row, column=4, value=r['cluster_id'])
    ws_detail.cell(row=row, column=5, value=r['문항'])
    ws_detail.cell(row=row, column=6, value=r.get('년도', ''))
    ws_detail.cell(row=row, column=7, value=r.get('회사명', ''))
    for c in range(1, 8):
        ws_detail.cell(row=row, column=c).border = thin_border
    row += 1
    if row % 2000 == 0:
        print(f'  처리 중... {row}행')

ws_detail.column_dimensions['A'].width = 8
ws_detail.column_dimensions['B'].width = 18
ws_detail.column_dimensions['C'].width = 18
ws_detail.column_dimensions['D'].width = 12
ws_detail.column_dimensions['E'].width = 80
ws_detail.column_dimensions['F'].width = 8
ws_detail.column_dimensions['G'].width = 15

print(f'Sheet 4: 전체문항상세 완료 ({row-2}개 문항)')

# 저장
output_path = 'results/전체_분류체계_검증용.xlsx'
wb.save(output_path)
print(f'\n저장 완료: {output_path}')
